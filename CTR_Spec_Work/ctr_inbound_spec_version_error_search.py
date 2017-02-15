'''
Created on Feb 15, 2017

@author: mstirling
'''

from openpyxl import load_workbook
#import re

in_folder = 'C:\Temp\python\in\CTR Specs/'.replace('\\','/')
in_filename = 'CTR_Mapping_Inbound.xlsx'

wb = load_workbook(in_folder + in_filename)
sheets =  wb.get_sheet_names()

for sheet in wb.worksheets:
    #print sheet.title + ', ' + str(sheet['D5'].value) 
    this_version = ''
    
    sheet_title_split = sheet.title.split('_')
    sheet_title_split_len = len(sheet_title_split)
    if sheet_title_split_len > 1:
        this_version = sheet_title_split[sheet_title_split_len-1]
        print sheet.title + ', ' + this_version
    else:
        print sheet.title + ', no_version'
    
    for i in range(1,500):
        if '1.8' in str(sheet['C' + str(i)].value):
            print sheet.title + ', ' + str(sheet['C' + str(i)].value)
        
        
    
#worksheet1 = wb2['Sheet1'] # one way to load a worksheet
#worksheet2 = wb2.get_sheet_by_name('Sheet2') # another way to load a worksheet
#print(worksheet1['D18'].value)

#for row in worksheet1.iter_rows():
#     print row[0].value()