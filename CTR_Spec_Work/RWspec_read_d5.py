'''
Created on Sep 16, 2016

@author: mstirling
'''

in_folder = 'C:/Users/mstirling/Desktop/'
in_filename = 'CTR_Mapping_Outbound_Riskwatch.xlsx'

from openpyxl import load_workbook
wb = load_workbook(in_folder + in_filename)
sheets =  wb.get_sheet_names()

for sheet in wb.worksheets:
    print sheet.title + ', ' + str(sheet['D5'].value) 

#worksheet1 = wb2['Sheet1'] # one way to load a worksheet
#worksheet2 = wb2.get_sheet_by_name('Sheet2') # another way to load a worksheet
#print(worksheet1['D18'].value)

#for row in worksheet1.iter_rows():
#     print row[0].value()